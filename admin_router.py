"""Interface d'administration — gestion des utilisateurs.

Réservée aux membres avec role='admin'.
"""

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from passlib.context import CryptContext

from heures_models import Membre, obtenir_session
from sqlalchemy.orm import Session
from sqlalchemy import create_engine, text
from pathlib import Path
from excel_writer import FICHIER_EXCEL
from parametres import designer_responsable_horaires, obtenir_responsable_horaires_id
from openpyxl import load_workbook

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
router = APIRouter()
templates = Jinja2Templates(directory="templates")


def _verifier_admin(request: Request, session: Session) -> bool:
    """Vérifie si le membre connecté est admin. Retourne False si non."""
    membre_id = request.session.get("membre_id")
    if not membre_id:
        return False
    membre = session.query(Membre).filter(Membre.id == membre_id).first()
    return membre is not None and membre.role == "admin"


@router.get("/admin/utilisateurs", response_class=HTMLResponse)
def lister_utilisateurs(
    request: Request,
    erreur: str | None = None,
    confirmation: str | None = None,
    session: Session = Depends(obtenir_session),
):
    if not _verifier_admin(request, session):
        return RedirectResponse(url="/login", status_code=303)

    membres = session.query(Membre).order_by(Membre.email).all()
    return templates.TemplateResponse(
        request,
        "admin/utilisateurs.html",
        {
            "membres": membres,
            "responsable_id": obtenir_responsable_horaires_id(),
            "erreur": erreur,
            "confirmation": confirmation,
            "actif": "admin_utilisateurs",
        },
    )


@router.post("/admin/utilisateurs/creer", response_class=HTMLResponse)
def creer_utilisateur(
    request: Request,
    email: str = Form(...),
    mot_de_passe: str = Form(...),
    nom: str = Form(""),
    role: str = Form("membre"),
    session: Session = Depends(obtenir_session),
):
    if not _verifier_admin(request, session):
        return RedirectResponse(url="/login", status_code=303)

    email = email.strip().lower()
    nom = nom.strip() or email.split("@")[0]
    erreur = None

    # Validation
    if not mot_de_passe or len(mot_de_passe) < 4:
        erreur = "Le mot de passe doit faire au moins 4 caractères."
    elif "@" not in email:
        erreur = "Adresse email invalide."
    elif role not in ("membre", "admin"):
        erreur = "Rôle invalide."
    else:
        existant = session.query(Membre).filter(Membre.email == email).first()
        if existant:
            erreur = f"Un compte existe déjà avec l'email {email}."

    if erreur:
        membres = session.query(Membre).order_by(Membre.email).all()
        return templates.TemplateResponse(
            request,
            "admin/utilisateurs.html",
            {
                "membres": membres,
                "erreur": erreur,
                "confirmation": None,
                "actif": "admin_utilisateurs",
            },
            status_code=400,
        )

    hash_pwd = pwd_context.hash(mot_de_passe)
    membre = Membre(
        nom=nom,
        email=email,
        mot_de_passe_hash=hash_pwd,
        role=role,
        actif=True,
    )
    session.add(membre)
    session.commit()

    return lister_utilisateurs(
        request,
        confirmation=f"Compte créé : {email} ({role})",
        session=session,
    )

@router.post("/admin/designer-responsable/{membre_id}", response_class=HTMLResponse)
def designer_responsable(request: Request, membre_id: int):
    session = next(obtenir_session())
    try:
        if not _verifier_admin(request, session):
            return RedirectResponse(url="/login", status_code=303)

        membre = session.query(Membre).filter(Membre.id == membre_id).first()
        if not membre:
            return RedirectResponse(
                url="/admin/utilisateurs",
                status_code=303,
            )

        current = obtenir_responsable_horaires_id()
        new_id = None if current == membre_id else membre_id
        designer_responsable_horaires(new_id)

        msg = f"Responsable horaires : {membre.nom} designe." if new_id else "Responsable horaires retire."
        return RedirectResponse(
            url=f"/admin/utilisateurs?confirmation={msg}",
            status_code=303,
        )
    finally:
        session.close()


@router.post("/admin/reset-donnees", response_class=HTMLResponse)
def reset_donnees_test(request: Request):
    """Supprime toutes les donnees de test (pointages, heures CAF,
    declarations, fiches de recuperation...) en conservant les membres,
    les taux horaires et les parametres."""
    session = next(obtenir_session())
    try:
        if not _verifier_admin(request, session):
            return RedirectResponse(url="/login", status_code=303)

        # 1. heures.db — supprimer donnees operationnelles
        session.execute(text("DELETE FROM validations_recup"))
        session.execute(text("DELETE FROM demandes_recup"))
        session.execute(text("DELETE FROM horaires"))
        session.execute(text("DELETE FROM jetons_connexion"))
        session.commit()

        # 2. bilan.db — supprimer heures_activite
        bd_bilan = Path(__file__).parent / "data" / "bilan.db"
        moteur_bilan = create_engine(f"sqlite:///{bd_bilan}")
        with moteur_bilan.connect() as conn:
            conn.execute(text("DELETE FROM heures_activite"))
            conn.commit()

        # 3. Fichier Excel pointages — vider les lignes de donnees
        if FICHIER_EXCEL.exists():
            classeur = load_workbook(FICHIER_EXCEL)
            feuille = classeur.active
            if feuille.max_row > 1:
                feuille.delete_rows(2, feuille.max_row - 1)
            classeur.save(FICHIER_EXCEL)
            classeur.close()

        return lister_utilisateurs(
            request,
            confirmation="Donnees de test reinitialisees (pointages, heures CAF, declarations). Membres et taux horaires conserves.",
            session=session,
        )
    finally:
        session.close()
