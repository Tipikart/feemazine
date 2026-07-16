"""Persistance des passages dans le fichier Excel.

Ce module est le seul point d'entree pour la lecture et l'ecriture des
passages. Les routes FastAPI l'appellent sans connaitre le detail du
stockage.

Chaque passage est soit anonyme (mode par defaut, compteurs manuels), soit
pseudonyme (mode carte, compteurs dupliques depuis la fiche carte). Dans les
deux cas, la ligne Excel ne contient aucune donnee d'identite directe.
remplacer_fichier() refuse tout fichier dont les colonnes ne correspondent
pas au format attendu.
"""

import io
import os
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

DATA_DIR = Path(__file__).parent / "data"
DOSSIER_SAUVEGARDES = DATA_DIR / "backups"
FICHIER_EXCEL = DATA_DIR / "passages.xlsx"
NOM_FEUILLE = "Passages"
EN_TETES_BASE = ["Date", "Heure", "Adultes", "Enfants", "Nouvelle famille"]
EN_TETES = EN_TETES_BASE + ["Mode", "Carte"]


class FichierVerrouille(Exception):
    """Le fichier Excel est ouvert dans une autre application."""


def chemin_fichier() -> Path:
    return FICHIER_EXCEL


def fichier_existe() -> bool:
    return FICHIER_EXCEL.exists()


def derniere_modification() -> datetime | None:
    if not FICHIER_EXCEL.exists():
        return None
    return datetime.fromtimestamp(FICHIER_EXCEL.stat().st_mtime)


def _ouvrir_ou_creer_classeur():
    if FICHIER_EXCEL.exists():
        classeur = load_workbook(FICHIER_EXCEL)
        feuille = classeur[NOM_FEUILLE]
        _migrer_en_tetes(feuille)
    else:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        classeur = Workbook()
        feuille = classeur.active
        feuille.title = NOM_FEUILLE
        feuille.append(EN_TETES)
    return classeur, feuille


def _migrer_en_tetes(feuille) -> None:
    """Ajoute les colonnes Mode et Carte si le fichier est au format 5 colonnes."""
    premiere_ligne = [feuille.cell(row=1, column=c).value for c in range(1, 8)]
    if premiere_ligne[:5] == EN_TETES_BASE and premiere_ligne[5] is None:
        feuille.cell(row=1, column=6, value="Mode")
        feuille.cell(row=1, column=7, value="Carte")


def _sauvegarder(classeur: Workbook) -> None:
    try:
        classeur.save(FICHIER_EXCEL)
    except PermissionError as erreur:
        raise FichierVerrouille(
            "Le fichier Excel est probablement ouvert dans une autre application."
        ) from erreur


def enregistrer_passage(
    adultes: int,
    enfants: int,
    nouvelle_famille: str,
    mode: str = "anonyme",
    carte_id: str | None = None,
) -> None:
    """Ajoute une ligne de passage au fichier Excel.

    En mode anonyme, les compteurs sont saisis manuellement.
    En mode carte, les compteurs sont dupliques depuis la fiche carte.
    """
    maintenant = datetime.now()
    classeur, feuille = _ouvrir_ou_creer_classeur()
    feuille.append(
        [
            maintenant.strftime("%Y-%m-%d"),
            maintenant.strftime("%H:%M:%S"),
            adultes,
            enfants,
            nouvelle_famille,
            mode,
            carte_id or "",
        ]
    )
    _sauvegarder(classeur)


def lire_tous_les_passages() -> list[dict]:
    if not FICHIER_EXCEL.exists():
        return []

    classeur = load_workbook(FICHIER_EXCEL, read_only=True)
    feuille = classeur[NOM_FEUILLE]
    passages = []
    for ligne in feuille.iter_rows(min_row=2, values_only=True):
        valeurs = list(ligne)
        if not valeurs or valeurs[0] is None:
            continue
        passages.append(
            {
                "date": valeurs[0],
                "heure": valeurs[1],
                "adultes": valeurs[2] or 0,
                "enfants": valeurs[3] or 0,
                "nouvelle_famille": valeurs[4] or "Non renseigne",
                "mode": valeurs[5] if len(valeurs) > 5 and valeurs[5] else "anonyme",
                "carte_id": valeurs[6] if len(valeurs) > 6 and valeurs[6] else None,
            }
        )
    classeur.close()
    return passages


def ouvrir_fichier() -> None:
    if not FICHIER_EXCEL.exists():
        raise FileNotFoundError("Aucun fichier de passages n'existe encore.")

    if hasattr(os, "startfile"):
        os.startfile(FICHIER_EXCEL)  # type: ignore[attr-defined]
    elif shutil.which("open"):
        os.system(f'open "{FICHIER_EXCEL}"')
    elif shutil.which("xdg-open"):
        os.system(f'xdg-open "{FICHIER_EXCEL}"')
    else:
        raise OSError("Aucune methode d'ouverture automatique disponible sur ce systeme.")


def remplacer_fichier(contenu_binaire: bytes) -> None:
    """Remplace le fichier de passages apres validation des colonnes.

    Accepte le format 5 colonnes (ancien) ou 7 colonnes (avec Mode/Carte).
    """
    try:
        classeur_importe = load_workbook(io.BytesIO(contenu_binaire))
    except Exception as erreur:
        raise ValueError("Le fichier envoye n'est pas un fichier Excel (.xlsx) valide.") from erreur

    if NOM_FEUILLE not in classeur_importe.sheetnames:
        raise ValueError(f"La feuille '{NOM_FEUILLE}' est introuvable dans le fichier importe.")

    feuille_importee = classeur_importe[NOM_FEUILLE]
    premiere_ligne = list(
        next(feuille_importee.iter_rows(min_row=1, max_row=1, values_only=True), ()) or ()
    )

    if premiere_ligne != EN_TETES and premiere_ligne != EN_TETES_BASE:
        raise ValueError(
            "Les en-tetes du fichier importe ne correspondent pas au format attendu : "
            + ", ".join(EN_TETES)
        )

    if FICHIER_EXCEL.exists():
        DOSSIER_SAUVEGARDES.mkdir(parents=True, exist_ok=True)
        horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(FICHIER_EXCEL, DOSSIER_SAUVEGARDES / f"passages_avant_import_{horodatage}.xlsx")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    try:
        classeur_importe.save(FICHIER_EXCEL)
    except PermissionError as erreur:
        raise FichierVerrouille(
            "Le fichier Excel est probablement ouvert dans une autre application."
        ) from erreur
